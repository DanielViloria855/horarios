# utils.py
import io
import calendar
from datetime import datetime, date, time, timedelta
from collections import defaultdict
from typing import Dict, Any

from flask import render_template_string
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# IMPORTAR db y modelos desde tu proyecto
from extensions import db
import models
Empleado = models.Empleado
Turno = models.Turno

# ---------------------------
# calcular_turno
# ---------------------------
def calcular_turno(fecha: date, entrada_time: time, salida_time: time) -> Dict[str, float]:
    entrada = datetime.combine(fecha, entrada_time)
    salida = datetime.combine(fecha, salida_time)
    if salida <= entrada:
        salida += timedelta(days=1)

    total_horas = (salida - entrada).total_seconds() / 3600.0

    corte_nocturno = datetime.combine(fecha, time(21, 0))  # 21:00
    medianoche = datetime.combine(fecha + timedelta(days=1), time(0, 0))

    fin_base = min(salida, medianoche)
    horas_dia_base = max(0.0, (fin_base - entrada).total_seconds() / 3600.0)

    nocturnas_base = 0.0
    if fin_base > corte_nocturno:
        nocturnas_base = max(0.0, (fin_base - max(entrada, corte_nocturno)).total_seconds() / 3600.0)

    horas_despues_medianoche = 0.0
    if salida > medianoche:
        horas_despues_medianoche = (salida - medianoche).total_seconds() / 3600.0

    return {
        "total_horas": round(total_horas, 2),
        "horas_dia_base": round(horas_dia_base, 2),
        "nocturnas_base": round(nocturnas_base, 2),
        "horas_despues_medianoche": round(horas_despues_medianoche, 2),
        "nocturnas_dia_siguiente": round(horas_despues_medianoche, 2),
    }

# ---------------------------
# preparar_datos_reporte - MODIFICADA
# ---------------------------
def preparar_datos_reporte(fecha_inicio: date, fecha_fin: date, tipo_datos: str = "totales") -> dict:
    """
    Devuelve datos con estructura mejorada:
    - 'dias': {date: {'total': float, 'hn': float, 'hn_dom': float}}
    - Asigna correctamente horas que cruzan medianoche al día siguiente
    - Calcula nocturnas normales y dominicales según día de la semana
    """
    consulta_inicio = fecha_inicio - timedelta(days=1)
    turnos = (
        db.session.query(Turno)
        .join(Empleado)
        .filter(Turno.fecha.between(consulta_inicio, fecha_fin))
        .order_by(Empleado.nombre)
        .all()
    )

    datos_empleados = defaultdict(lambda: {
        "nombre": "", "cargo": "",
        "dias": defaultdict(lambda: {"total": 0.0, "hn": 0.0, "hn_dom": 0.0}),
        "total_nocturnas": 0.0, 
        "total_nocturnas_dominicales": 0.0,
        "total_extras": 0.0, 
        "total_horas_normales": 0.0
    })

    for t in turnos:
        if not getattr(t, "hora_entrada", None):
            continue
            
        emp_id = t.id_empleado
        if not datos_empleados[emp_id]["nombre"]:
            datos_empleados[emp_id]["nombre"] = t.empleado.nombre
            datos_empleados[emp_id]["cargo"] = getattr(t.empleado, "cargo", "")

        res = calcular_turno(t.fecha, t.hora_entrada, t.hora_salida)
        total_horas_turno = res["total_horas"]
        hn_base = res["nocturnas_base"]  # Nocturnas antes de medianoche
        hn_next = res["nocturnas_dia_siguiente"]  # Nocturnas después de medianoche
        horas_despues_medianoche = res["horas_despues_medianoche"]

        # Determinar si los días son domingos
        es_domingo_turno = t.fecha.weekday() == 6
        dia_siguiente = t.fecha + timedelta(days=1)
        es_domingo_siguiente = dia_siguiente.weekday() == 6

        # Inicializar contadores por tipo de hora
        hn_on_fecha = 0.0
        hn_dom_on_fecha = 0.0
        hn_on_next = 0.0
        hn_dom_on_next = 0.0

        # Asignar nocturnas antes de medianoche (pertenecen a t.fecha)
        if hn_base > 0:
            if es_domingo_turno:
                hn_dom_on_fecha += hn_base
            else:
                hn_on_fecha += hn_base

        # Asignar nocturnas después de medianoche (pertenecen a dia_siguiente)
        if hn_next > 0:
            if es_domingo_siguiente:
                hn_dom_on_next += hn_next
            else:
                hn_on_next += hn_next

        # Calcular horas normales (total - todas las nocturnas)
        horas_normales = total_horas_turno - (hn_on_fecha + hn_dom_on_fecha + hn_on_next + hn_dom_on_next)

        # PARTE DEL TURNO EN EL DÍA ACTUAL (antes de medianoche)
        if fecha_inicio <= t.fecha <= fecha_fin:
            # Asignar horas del día base
            datos_empleados[emp_id]["dias"][t.fecha]["total"] += res["horas_dia_base"]
            datos_empleados[emp_id]["dias"][t.fecha]["hn"] += hn_on_fecha
            datos_empleados[emp_id]["dias"][t.fecha]["hn_dom"] += hn_dom_on_fecha

            # Acumular totales (se suman todas las nocturnas del turno)
            datos_empleados[emp_id]["total_nocturnas"] += hn_on_fecha + hn_on_next
            datos_empleados[emp_id]["total_nocturnas_dominicales"] += hn_dom_on_fecha + hn_dom_on_next
            datos_empleados[emp_id]["total_horas_normales"] += horas_normales
            datos_empleados[emp_id]["total_extras"] += max(0, res["horas_dia_base"] - 8)

        # PARTE DEL TURNO EN EL DÍA SIGUIENTE (después de medianoche)
        if horas_despues_medianoche > 0 and (fecha_inicio <= dia_siguiente <= fecha_fin):
            # Asignar horas del día siguiente
            datos_empleados[emp_id]["dias"][dia_siguiente]["total"] += res["horas_despues_medianoche"]
            datos_empleados[emp_id]["dias"][dia_siguiente]["hn"] += hn_on_next
            datos_empleados[emp_id]["dias"][dia_siguiente]["hn_dom"] += hn_dom_on_next

            # Los totales ya se sumaron arriba para evitar duplicación
            datos_empleados[emp_id]["total_extras"] += max(0, res["horas_despues_medianoche"] - 8)

    return dict(datos_empleados)

# ---------------------------
# estructurar_reporte - COMPATIBILIDAD
# ---------------------------
def estructurar_reporte(datos: dict, fecha_inicio: date, fecha_fin: date, periodo_tipo: str, tipo_datos: str = "totales") -> dict:
    """
    Versión compatibilizada que trabaja con la nueva estructura de 'dias'
    """
    dias_semana = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
    meses_es = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
        7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    mes_nombre = meses_es.get(fecha_inicio.month, fecha_inicio.strftime("%B").upper())
    
    if periodo_tipo == "personalizado":
        titulo = f"REPORTE PERSONALIZADO {fecha_inicio.day}/{fecha_inicio.month} - {fecha_fin.day}/{fecha_fin.month} DE {fecha_inicio.year}"
    else:
        titulo = f"REPORTE {mes_nombre} {fecha_inicio.day} - {fecha_fin.day} DE {fecha_inicio.year}"

    todos_los_dias = []
    current_date = fecha_inicio
    while current_date <= fecha_fin:
        todos_los_dias.append(current_date)
        current_date += timedelta(days=1)

    # Determinar si necesitamos dividir en semanas para quincenal y mensual
    dividir_en_semanas = periodo_tipo in ["quincenal", "mensual"]
    
    if dividir_en_semanas:
        # Dividir en semanas (lunes a domingo)
        semanas = []
        semana_actual = []
        for dia in todos_los_dias:
            if not semana_actual:
                semana_actual.append(dia)
            else:
                if dia.weekday() == 0:  # Lunes - nueva semana
                    semanas.append(semana_actual)
                    semana_actual = [dia]
                else:
                    semana_actual.append(dia)
        if semana_actual:
            semanas.append(semana_actual)
        
        # Reconstruir la lista de días con separadores de semana
        todos_los_dias_con_separadores = []
        for i, semana in enumerate(semanas):
            todos_los_dias_con_separadores.extend(semana)
            if i < len(semanas) - 1:
                # Agregar un día separador
                todos_los_dias_con_separadores.append("SEPARADOR")
    else:
        todos_los_dias_con_separadores = todos_los_dias

    n = len(todos_los_dias)
    split_index = n // 2
    first_block = todos_los_dias[:split_index]
    second_block = todos_los_dias[split_index:]

    columnas = [{"key": "employee", "label": "EMPLEAD@"}]

    for dia in first_block:
        if dia == "SEPARADOR":
            columnas.append({
                "key": "separador",
                "label_day": "",
                "label_date": "",
                "is_separador": True
            })
        else:
            columnas.append({
                "key": f"day_{dia.strftime('%Y-%m-%d')}",
                "label_day": dias_semana[dia.weekday()],
                "label_date": dia.day,
                "block": 1
            })

    columnas.append({
        "key": "subtotal_block1",
        "label_day": None,
        "label_date": "S1",
        "is_subtotal": True,
        "block": "subtotal"
    })

    for dia in second_block:
        if dia == "SEPARADOR":
            columnas.append({
                "key": "separador",
                "label_day": "",
                "label_date": "",
                "is_separador": True
            })
        else:
            columnas.append({
                "key": f"day_{dia.strftime('%Y-%m-%d')}",
                "label_day": dias_semana[dia.weekday()],
                "label_date": dia.day,
                "block": 2
            })

    columnas.extend([
        {"key": "hn", "label": "H.N."},
        {"key": "hn_dom", "label": "H.N.Dom"},
        {"key": "total", "label": "TOTAL"}
    ])

    filas = []
    for emp_id, emp_data in datos.items():
        # Calcular total según tipo de datos
        if tipo_datos == "nocturnas":
            total_horas = emp_data['total_nocturnas'] + emp_data['total_nocturnas_dominicales']
        else:
            total_horas = sum(dia_data['total'] for dia_data in emp_data['dias'].values())
            
        celdas = {}
        for dia in todos_los_dias_con_separadores:
            if dia == "SEPARADOR":
                key = f"separador"
                celdas[key] = ""
            else:
                key = f"day_{dia.strftime('%Y-%m-%d')}"
                dia_info = emp_data['dias'].get(dia, {"total": 0.0, "hn": 0.0, "hn_dom": 0.0})
                
                if tipo_datos == "nocturnas":
                    valor_dia = dia_info["hn"] + dia_info["hn_dom"]
                    celdas[key] = f"{valor_dia:.1f}" if valor_dia > 0 else "-"
                else:
                    valor_dia = dia_info["total"]
                    celdas[key] = f"{valor_dia:.1f}" if valor_dia > 0 else ""

        # Calcular subtotales por bloques
        subtotal_block1 = 0.0
        for d in first_block:
            if d != "SEPARADOR":
                dia_info = emp_data['dias'].get(d, {"total": 0.0, "hn": 0.0, "hn_dom": 0.0})
                if tipo_datos == "nocturnas":
                    subtotal_block1 += dia_info["hn"] + dia_info["hn_dom"]
                else:
                    subtotal_block1 += dia_info["total"]
                    
        celdas["subtotal_block1"] = f"{subtotal_block1:.1f}" if subtotal_block1 > 0 else ""

        # Totales generales
        celdas["hn"] = f"{emp_data['total_nocturnas']:.1f}"
        celdas["hn_dom"] = f"{emp_data['total_nocturnas_dominicales']:.1f}"
        celdas["total"] = f"{total_horas:.1f}"

        fila = {
            "employee_id": emp_id,
            "employee_name": emp_data['nombre'],
            "cells": celdas
        }
        filas.append(fila)

    return {
        "report": {
            "title": titulo,
            "period": {
                "type": periodo_tipo,
                "start_date": fecha_inicio.strftime('%Y-%m-%d'),
                "end_date": fecha_fin.strftime('%Y-%m-%d'),
                "split_index": split_index
            },
            "columns": columnas,
            "rows": filas
        }
    }

# ---------------------------
# estructurar_reporte_con_columnas - MODIFICADA PRINCIPAL
# ---------------------------
def estructurar_reporte_con_columnas(datos: dict, fecha_inicio: date, fecha_fin: date, periodo_tipo: str, tipo_datos: str = "totales") -> dict:
    """
    Estructura el reporte con columnas específicas:
    - Para quincenal: columnas H.N, H.N.DOM y TOTAL por semana
    - Para mensual: dividido en dos quincenas
    - Aplica regla 7/1: si hay 7h nocturnas totales con 1h dominical -> mostrar H.N=6, H.N.DOM=1
    """
    dias_semana = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
    meses_es = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
        7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    mes_nombre = meses_es.get(fecha_inicio.month, fecha_inicio.strftime("%B").upper())
    
    # Construir título según tipo de período
    if periodo_tipo == "personalizado":
        titulo = f"REPORTE PERSONALIZADO {fecha_inicio.strftime('%d/%m')} - {fecha_fin.strftime('%d/%m')} DE {fecha_inicio.year}"
    elif periodo_tipo == "semanal":
        titulo = f"REPORTE SEMANAL {fecha_inicio.strftime('%d/%m')} - {fecha_fin.strftime('%d/%m')} DE {fecha_inicio.year}"
    elif periodo_tipo == "quincenal":
        quincena = "PRIMERA" if fecha_inicio.day == 1 else "SEGUNDA"
        titulo = f"REPORTE QUINCENAL {quincena} QUINCENA {mes_nombre} {fecha_inicio.year}"
    else:  # mensual
        titulo = f"REPORTE MENSUAL {mes_nombre} {fecha_inicio.year}"

    # Obtener todos los días del rango
    todos_los_dias = []
    current_date = fecha_inicio
    while current_date <= fecha_fin:
        todos_los_dias.append(current_date)
        current_date += timedelta(days=1)

    # Dividir en bloques según tipo de período
    if periodo_tipo == "quincenal":
        # Quincenal: dividir en 2 semanas
        semana1 = todos_los_dias[:7]  # Primera semana (días 1-7)
        semana2 = todos_los_dias[7:15] if len(todos_los_dias) > 7 else []  # Segunda semana (días 8-15)
        bloques = [semana1, semana2] if semana2 else [semana1]
        
    elif periodo_tipo == "mensual":
        # Mensual: dividir en 2 quincenas
        quincena1 = [d for d in todos_los_dias if d.day <= 15]
        quincena2 = [d for d in todos_los_dias if d.day > 15]
        bloques = [quincena1, quincena2] if quincena2 else [quincena1]
        
    else:
        # Semanal o personalizado: un solo bloque
        bloques = [todos_los_dias]

    # Construir columnas
    columnas = [{"key": "employee", "label": "EMPLEADO", "tipo": "nombre"}]
    
    # Agregar columnas de días por cada bloque
    for i, bloque in enumerate(bloques):
        for dia in bloque:
            columnas.append({
                "key": f"day_{dia.strftime('%Y-%m-%d')}",
                "label_day": dias_semana[dia.weekday()],
                "label_date": dia.day,
                "block": i+1,
                "tipo": "dia"
            })
        
        # Agregar columnas específicas para cada bloque según el tipo de período
        if periodo_tipo == "quincenal":
            columnas.extend([
                {"key": f"hn_semana_{i+1}", "label": "H.N", "is_semanal": True, "block": i+1, "tipo": "hn"},
                {"key": f"hn_dom_semana_{i+1}", "label": "H.N.DOM", "is_semanal": True, "block": i+1, "tipo": "hndom"},
                {"key": f"total_semana_{i+1}", "label": "TOTAL", "is_semanal": True, "block": i+1, "tipo": "total_semana"}
            ])
        elif periodo_tipo == "mensual":
            columnas.extend([
                {"key": f"hn_quincena_{i+1}", "label": "H.N", "is_quincenal": True, "block": i+1, "tipo": "hn"},
                {"key": f"hn_dom_quincena_{i+1}", "label": "H.N.DOM", "is_quincenal": True, "block": i+1, "tipo": "hndom"},
                {"key": f"total_quincena_{i+1}", "label": "TOTAL", "is_quincenal": True, "block": i+1, "tipo": "total_quincena"}
            ])

    # Agregar totales generales al final
    columnas.extend([
        {"key": "hn_total", "label": "H.N TOTAL", "tipo": "hn_total"},
        {"key": "hn_dom_total", "label": "H.N.DOM TOTAL", "tipo": "hndom_total"},
        {"key": "total_general", "label": "TOTAL GENERAL", "tipo": "total_general"}
    ])

    # Preparar filas con datos
    filas = []
    for emp_id, emp_data in datos.items():
        celdas = {}
        
        # Llenar datos por día
        for dia in todos_los_dias:
            key = f"day_{dia.strftime('%Y-%m-%d')}"
            dia_info = emp_data['dias'].get(dia, {"total": 0.0, "hn": 0.0, "hn_dom": 0.0})
            
            if tipo_datos == "nocturnas":
                # Mostrar solo horas nocturnas (hn + hn_dom)
                valor_dia = dia_info["hn"] + dia_info["hn_dom"]
                celdas[key] = f"{valor_dia:.1f}" if valor_dia > 0 else "-"
            else:
                # Mostrar horas totales
                valor_dia = dia_info["total"]
                celdas[key] = f"{valor_dia:.1f}" if valor_dia > 0 else ""

        # Calcular subtotales por bloque y aplicar regla 7/1
        for i, bloque in enumerate(bloques):
            # Calcular horas por bloque según tipo de datos
            horas_bloque = 0.0
            hn_bloque = 0.0
            hn_dom_bloque = 0.0
            
            for dia in bloque:
                dia_info = emp_data['dias'].get(dia, {"total": 0.0, "hn": 0.0, "hn_dom": 0.0})
                hn_bloque += dia_info["hn"]
                hn_dom_bloque += dia_info["hn_dom"]
                if tipo_datos == "nocturnas":
                    horas_bloque += dia_info["hn"] + dia_info["hn_dom"]
                else:
                    horas_bloque += dia_info["total"]

            # APLICAR REGLA 7/1 - Si hay aproximadamente 7h nocturnas con al menos 1h dominical
            total_nocturnas_bloque = hn_bloque + hn_dom_bloque
            if abs(total_nocturnas_bloque - 7.0) < 0.1:  # Tolerancia para floats
                if hn_bloque >= 6 and hn_dom_bloque >= 1:
                    # Caso: 6 normales + 1 dominical -> mostrar H.N=6, H.N.DOM=1
                    hn_bloque_display = 6.0
                    hn_dom_bloque_display = 1.0
                elif hn_dom_bloque >= 6 and hn_bloque >= 1:
                    # Caso espejo: 1 normal + 6 dominicales -> mostrar H.N=1, H.N.DOM=6
                    hn_bloque_display = 1.0
                    hn_dom_bloque_display = 6.0
                else:
                    # No aplicar ajuste
                    hn_bloque_display = hn_bloque
                    hn_dom_bloque_display = hn_dom_bloque
            else:
                # No aplicar ajuste
                hn_bloque_display = hn_bloque
                hn_dom_bloque_display = hn_dom_bloque

            # Para reportes nocturnos, el TOTAL del bloque debe ser solo nocturnas
            if tipo_datos == "nocturnas":
                horas_bloque = total_nocturnas_bloque

            # Asignar valores a las celdas del bloque
            if periodo_tipo == "quincenal":
                celdas[f"hn_semana_{i+1}"] = f"{hn_bloque_display:.1f}" if hn_bloque_display > 0 else "0.0"
                celdas[f"hn_dom_semana_{i+1}"] = f"{hn_dom_bloque_display:.1f}" if hn_dom_bloque_display > 0 else "0.0"
                celdas[f"total_semana_{i+1}"] = f"{horas_bloque:.1f}" if horas_bloque > 0 else "0.0"
            elif periodo_tipo == "mensual":
                celdas[f"hn_quincena_{i+1}"] = f"{hn_bloque_display:.1f}" if hn_bloque_display > 0 else "0.0"
                celdas[f"hn_dom_quincena_{i+1}"] = f"{hn_dom_bloque_display:.1f}" if hn_dom_bloque_display > 0 else "0.0"
                celdas[f"total_quincena_{i+1}"] = f"{horas_bloque:.1f}" if horas_bloque > 0 else "0.0"

        # Totales generales
        if tipo_datos == "nocturnas":
            total_general = emp_data['total_nocturnas'] + emp_data['total_nocturnas_dominicales']
        else:
            total_general = sum(dia_data['total'] for dia_data in emp_data['dias'].values())
            
        total_hn = emp_data['total_nocturnas']
        total_hn_dom = emp_data['total_nocturnas_dominicales']
        
        celdas["hn_total"] = f"{total_hn:.1f}"
        celdas["hn_dom_total"] = f"{total_hn_dom:.1f}"
        celdas["total_general"] = f"{total_general:.1f}"

        fila = {
            "employee_id": emp_id,
            "employee_name": emp_data['nombre'],
            "cells": celdas
        }
        filas.append(fila)

    return {
        "report": {
            "title": titulo,
            "period": {
                "type": periodo_tipo,
                "start_date": fecha_inicio.strftime('%Y-%m-%d'),
                "end_date": fecha_fin.strftime('%Y-%m-%d'),
            },
            "columns": columnas,
            "rows": filas,
            "bloques": len(bloques)
        }
    }

# ---------------------------
# FUNCIONES DE GENERACIÓN (sin cambios)
# ---------------------------
def generar_pdf_tabla(datos_estructurados: dict) -> io.BytesIO:
    """Genera PDF con layout fijo (A4 landscape)"""
    reporte = datos_estructurados["report"]
    columns = reporte["columns"]
    rows = reporte["rows"]

    css = """
    @page { size: A4 landscape; margin: 10mm; }
    body { font-family: Arial, Helvetica, sans-serif; font-size: 9px; margin: 0; padding: 0; }
    .title { text-align: center; font-size: 14px; font-weight: bold; margin: 6px 0 8px 0; }
    table { width: 100%; border-collapse: collapse; table-layout: fixed; }
    th, td { border: 1px solid #000; padding: 4px; vertical-align: middle; box-sizing: border-box; }
    th { background: #f5f5f5; font-weight: bold; text-align: center; }
    .col-employee { width: 220px; min-width: 220px; max-width: 220px; text-align: left; padding-left: 6px; word-wrap: break-word; white-space: normal; }
    .col-day { width: 30px; min-width: 30px; max-width: 30px; }
    .col-separador { width: 10px; min-width: 10px; max-width: 10px; background:#ddd; }
    .col-subtotal { width: 48px; min-width: 48px; max-width: 48px; background:#efefef; }
    .col-total { width: 70px; min-width: 70px; max-width: 70px; background:#e9e9e9; }
    thead th.small { font-size: 8px; padding: 2px; }
    td.num, th.num { text-align: center; white-space: nowrap; }
    td.employee { word-break: break-word; white-space: normal; }
    td { max-height: 200px; overflow: hidden; }
    """

    html_template = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <style>{css}</style>
    </head>
    <body>
      <div class="title">{reporte["title"]}</div>
      <table>
        <thead>
          <tr>
            <th class="col-employee" rowspan="2">{columns[0]['label']}</th>
            {{% for col in columns[1:] %}}
                {{% if col.get('is_separador') %}}
                    <th class="col-separador" rowspan="2"></th>
                {{% elif col.get('is_subtotal') %}}
                    <th class="col-subtotal" rowspan="2">{{{{ col.label_date }}}}</th>
                {{% elif col.key.startswith('day_') %}}
                    <th class="col-day">{{{{ col.label_day }}}}</th>
                {{% else %}}
                    <th class="col-total" rowspan="2">{{{{ col.label }}}}</th>
                {{% endif %}}
            {{% endfor %}}
          </tr>
          <tr>
            {{% for col in columns[1:] %}}
                {{% if col.key.startswith('day_') %}}
                    <th class="num small">{{{{ col.label_date }}}}</th>
                {{% endif %}}
            {{% endfor %}}
          </tr>
        </thead>
        <tbody>
          {{% for row in rows %}}
            <tr>
              <td class="employee">{{{{ row.employee_name }}}}</td>
              {{% for col in columns[1:] %}}
                {{% if col.get('is_separador') %}}
                    <td class="col-separador"></td>
                {{% elif col.get('is_subtotal') %}}
                    <td class="num col-subtotal">{{{{ row.cells.get(col.key, '') }}}}</td>
                {{% elif col.key.startswith('day_') %}}
                    <td class="num col-day">{{{{ row.cells.get(col.key, '') }}}}</td>
                {{% else %}}
                    <td class="num col-total">{{{{ row.cells.get(col.key, '') }}}}</td>
                {{% endif %}}
              {{% endfor %}}
            </tr>
          {{% endfor %}}
        </tbody>
      </table>
    </body>
    </html>
    """

    html_content = render_template_string(html_template,
                                          columns=columns,
                                          rows=rows,
                                          title=reporte["title"])
    buffer = io.BytesIO()
    pisa_status = pisa.CreatePDF(src=html_content, dest=buffer)
    if pisa_status.err:
        raise RuntimeError("Error generando PDF con xhtml2pdf")
    buffer.seek(0)
    return buffer

def generar_excel_tabla(datos_estructurados: dict) -> io.BytesIO:
    reporte = datos_estructurados["report"]
    columnas = reporte["columns"]
    filas = reporte["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    header_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    total_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    subtotal_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    separador_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    total_columns = len(columnas)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1, value=reporte["title"])
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center

    row_header_1 = 2
    row_header_2 = 3
    col_idx = 1
    for col in columnas:
        if col["key"] == "employee":
            cell = ws.cell(row=row_header_1, column=col_idx, value=col["label"])
            cell.font = header_font
            cell.alignment = center
            cell.fill = total_fill
            cell.border = border
            ws.merge_cells(start_row=row_header_1, start_column=col_idx, end_row=row_header_2, end_column=col_idx)
            col_idx += 1
        elif col.get("is_separador"):
            cell = ws.cell(row=row_header_1, column=col_idx, value="")
            cell.fill = separador_fill
            cell.border = border
            ws.merge_cells(start_row=row_header_1, start_column=col_idx, end_row=row_header_2, end_column=col_idx)
            col_idx += 1
        elif col.get("is_subtotal"):
            cell = ws.cell(row=row_header_1, column=col_idx, value=col["label_date"])
            cell.font = header_font
            cell.alignment = center
            cell.fill = subtotal_fill
            cell.border = border
            ws.merge_cells(start_row=row_header_1, start_column=col_idx, end_row=row_header_2, end_column=col_idx)
            col_idx += 1
        elif col["key"].startswith("day_"):
            cell = ws.cell(row=row_header_1, column=col_idx, value=col["label_day"])
            cell.font = header_font
            cell.alignment = center
            cell.fill = total_fill
            cell.border = border
            cell2 = ws.cell(row=row_header_2, column=col_idx, value=col["label_date"])
            cell2.font = header_font
            cell2.alignment = center
            cell2.fill = total_fill
            cell2.border = border
            col_idx += 1
        else:
            cell = ws.cell(row=row_header_1, column=col_idx, value=col["label"])
            cell.font = header_font
            cell.alignment = center
            cell.fill = total_fill
            cell.border = border
            ws.merge_cells(start_row=row_header_1, start_column=col_idx, end_row=row_header_2, end_column=col_idx)
            col_idx += 1

    # datos
    row_num = 4
    for row in filas:
        col_idx = 1
        for col in columnas:
            cell = ws.cell(row=row_num, column=col_idx)
            if col["key"] == "employee":
                cell.value = row["employee_name"]
                cell.alignment = left
            else:
                cell.value = row["cells"].get(col["key"], "")
                cell.alignment = center
            cell.border = border
            if col.get("is_separador"):
                cell.fill = separador_fill
            elif col.get("is_subtotal"):
                cell.fill = subtotal_fill
            elif col["key"] in ["hn", "hn_dom", "total"]:
                cell.fill = total_fill
            col_idx += 1
        row_num += 1

    # ajustar anchos
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 20)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ---------------------------
# Compatibilidad: wrapper de generacion detallada
# ---------------------------
def generar_pdf_detallado(datos: dict, fecha_inicio: date, fecha_fin: date) -> io.BytesIO:
    datos_estructurados = estructurar_reporte(datos, fecha_inicio, fecha_fin, "personalizado")
    return generar_pdf_tabla(datos_estructurados)

def generar_excel_detallado(datos: dict, fecha_inicio: date, fecha_fin: date) -> io.BytesIO:
    datos_estructurados = estructurar_reporte(datos, fecha_inicio, fecha_fin, "personalizado")
    return generar_excel_tabla(datos_estructurados)

def generar_pdf_con_columnas(datos_estructurados: dict) -> io.BytesIO:
    """Genera PDF con las nuevas columnas específicas"""
    reporte = datos_estructurados["report"]
    columns = reporte["columns"]
    rows = reporte["rows"]

    css = """
    @page { size: A4 landscape; margin: 10mm; }
    body { font-family: Arial, Helvetica, sans-serif; font-size: 9px; margin: 0; padding: 0; }
    .title { text-align: center; font-size: 14px; font-weight: bold; margin: 6px 0 8px 0; }
    table { width: 100%; border-collapse: collapse; table-layout: fixed; }
    th, td { border: 1px solid #000; padding: 4px; vertical-align: middle; box-sizing: border-box; }
    th { background: #f5f5f5; font-weight: bold; text-align: center; }
    .col-employee { width: 180px; min-width: 180px; max-width: 180px; text-align: left; padding-left: 6px; }
    .col-day { width: 25px; min-width: 25px; max-width: 25px; }
    .col-hn { width: 35px; min-width: 35px; max-width: 35px; background: #e8f4fd; }
    .col-hn-dom { width: 45px; min-width: 45px; max-width: 45px; background: #e8f4fd; }
    .col-total-block { width: 40px; min-width: 40px; max-width: 40px; background: #f0f0f0; }
    .col-total-general { width: 50px; min-width: 50px; max-width: 50px; background: #d4edda; font-weight: bold; }
    .col-separator { width: 5px; min-width: 5px; max-width: 5px; background: #ddd; }
    thead th.small { font-size: 8px; padding: 2px; }
    td.num, th.num { text-align: center; white-space: nowrap; }
    td.employee { word-break: break-word; white-space: normal; }
    """

    html_template = """
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <style>{{ css }}</style>
    </head>
    <body>
      <div class="title">{{ title }}</div>
      <table>
        <thead>
          <tr>
            {% for col in columns %}
              {% if col.key == 'employee' %}
                <th class="col-employee" rowspan="2">{{ col.label }}</th>
              {% elif col.key.startswith('day_') %}
                <th class="col-day">{{ col.label_day }}</th>
              {% elif col.get('is_semanal') or col.get('is_quincenal') %}
                <th class="{% if 'hn' in col.key %}col-hn{% elif 'dom' in col.key %}col-hn-dom{% else %}col-total-block{% endif %}" rowspan="2">{{ col.label }}</th>
              {% elif col.key in ['hn_total', 'hn_dom_total'] %}
                <th class="col-hn" rowspan="2">{{ col.label }}</th>
              {% elif col.key == 'total_general' %}
                <th class="col-total-general" rowspan="2">{{ col.label }}</th>
              {% endif %}
            {% endfor %}
          </tr>
          <tr>
            {% for col in columns %}
              {% if col.key.startswith('day_') %}
                <th class="num small">{{ col.label_date }}</th>
              {% endif %}
            {% endfor %}
          </tr>
        </thead>
        <tbody>
          {% for row in rows %}
            <tr>
              {% for col in columns %}
                {% if col.key == 'employee' %}
                  <td class="employee">{{ row.employee_name }}</td>
                {% elif col.key.startswith('day_') or col.get('is_semanal') or col.get('is_quincenal') or col.key in ['hn_total', 'hn_dom_total', 'total_general'] %}
                  <td class="num {% if 'hn' in col.key %}col-hn{% elif 'dom' in col.key %}col-hn-dom{% elif 'total' in col.key %}col-total-block{% else %}col-day{% endif %}">
                    {{ row.cells.get(col.key, '') }}
                  </td>
                {% endif %}
              {% endfor %}
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </body>
    </html>
    """

    html_content = render_template_string(html_template,
                                         css=css,
                                         columns=columns,
                                         rows=rows,
                                         title=reporte["title"])
    
    buffer = io.BytesIO()
    pisa_status = pisa.CreatePDF(src=html_content, dest=buffer)
    if pisa_status.err:
        raise RuntimeError("Error generando PDF con xhtml2pdf")
    buffer.seek(0)
    return buffer

def generar_excel_con_columnas(datos_estructurados: dict) -> io.BytesIO:
    """Genera Excel con las nuevas columnas específicas"""
    reporte = datos_estructurados["report"]
    columnas = reporte["columns"]
    filas = reporte["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    header_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    hn_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    hn_dom_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    total_block_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    total_general_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    total_columns = len(columnas)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1, value=reporte["title"])
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center

    row_header_1 = 2
    row_header_2 = 3
    col_idx = 1

    for col in columnas:
        if col["key"] == "employee":
            cell = ws.cell(row=row_header_1, column=col_idx, value=col["label"])
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            ws.merge_cells(start_row=row_header_1, start_column=col_idx, end_row=row_header_2, end_column=col_idx)
            col_idx += 1
            
        elif col["key"].startswith("day_"):
            cell1 = ws.cell(row=row_header_1, column=col_idx, value=col["label_day"])
            cell1.font = header_font
            cell1.alignment = center
            cell1.border = border
            cell2 = ws.cell(row=row_header_2, column=col_idx, value=col["label_date"])
            cell2.font = header_font
            cell2.alignment = center
            cell2.border = border
            col_idx += 1
            
        elif col.get("is_semanal") or col.get("is_quincenal"):
            cell = ws.cell(row=row_header_1, column=col_idx, value=col["label"])
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            
            if "hn" in col["key"] and "dom" not in col["key"]:
                cell.fill = hn_fill
            elif "hn_dom" in col["key"]:
                cell.fill = hn_dom_fill
            else:
                cell.fill = total_block_fill
                
            ws.merge_cells(start_row=row_header_1, start_column=col_idx, end_row=row_header_2, end_column=col_idx)
            col_idx += 1
            
        else:
            cell = ws.cell(row=row_header_1, column=col_idx, value=col["label"])
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            
            if col["key"] in ["hn_total", "hn_dom_total"]:
                cell.fill = hn_fill
            elif col["key"] == "total_general":
                cell.fill = total_general_fill
                
            ws.merge_cells(start_row=row_header_1, start_column=col_idx, end_row=row_header_2, end_column=col_idx)
            col_idx += 1

    # datos
    row_num = 4
    for row in filas:
        col_idx = 1
        for col in columnas:
            cell = ws.cell(row=row_num, column=col_idx)
            
            if col["key"] == "employee":
                cell.value = row["employee_name"]
                cell.alignment = left
            else:
                cell.value = row["cells"].get(col["key"], "")
                cell.alignment = center
                
            cell.border = border
            
            if col["key"] in ["hn_total", "hn_dom_total"]:
                cell.fill = hn_fill
            elif col["key"] == "total_general":
                cell.fill = total_general_fill
            elif col.get("is_semanal") or col.get("is_quincenal"):
                if "hn" in col["key"] and "dom" not in col["key"]:
                    cell.fill = hn_fill
                elif "hn_dom" in col["key"]:
                    cell.fill = hn_dom_fill
                else:
                    cell.fill = total_block_fill
                    
            col_idx += 1
        row_num += 1

    # ajustar anchos
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer